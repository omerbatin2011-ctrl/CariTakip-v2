import os
from datetime import datetime

from PySide6.QtCore import QDate, Qt, QTimer
from PySide6.QtGui import QPageSize, QTextDocument
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import (
    QDateEdit,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
)

from core.config import BASE_DIR
from moduller.db import baglan, db_baglan, urun_tablolari_olustur
from moduller.sistem import firma_bilgisi_getir
from moduller.yardimci import para_yaz

try:
    from kur_modulu import tcmb_usd_kuru_al, tl_karsiligi_hesapla
except Exception:
    tcmb_usd_kuru_al = None
    def tl_karsiligi_hesapla(tutar, para_birimi="TL", kur=1):
        try:
            tutar = float(tutar or 0)
            kur = float(kur or 1)
        except Exception:
            return 0.0
        return tutar if str(para_birimi).upper() == "TL" else tutar * kur

class AnaPencereRaporMixin:
    def kart_olustur(self, baslik, deger, ikon=""):
        kutu = QFrame()
        kutu.setObjectName("InfoCard")
        layout = QVBoxLayout()
        layout.setContentsMargins(18, 14, 18, 14)

        ust = QLabel(f"{ikon}  {baslik}" if ikon else baslik)
        ust.setStyleSheet("font-size:14px;color:#6B7280;font-weight:bold;")

        lblDeger = QLabel(deger)
        lblDeger.setStyleSheet("font-size:26px;font-weight:bold;color:#0D47A1;padding-top:4px;")

        layout.addWidget(ust)
        layout.addWidget(lblDeger)

        kutu.setLayout(layout)
        self.kart_layout.addWidget(kutu)

        return lblDeger

    def ozet_yukle(self):
        conn = baglan()
        cur = conn.cursor()

        cur.execute("SELECT COUNT(*) FROM cariler WHERE COALESCE(aktif,1)=1")
        toplam_cari = cur.fetchone()[0] or 0

        cur.execute("SELECT SUM(tutar) FROM hareketler WHERE tip='BORÇ'")
        toplam_borc = cur.fetchone()[0] or 0

        cur.execute("SELECT SUM(tutar) FROM hareketler WHERE tip='TAHSİLAT'")
        toplam_tahsilat = cur.fetchone()[0] or 0

        conn.close()

        kalan = toplam_borc - toplam_tahsilat

        self.lblToplamCari.setText(str(toplam_cari))
        self.lblToplamBorc.setText(para_yaz(float(kalan)))
        self.lblToplamTahsilat.setText(para_yaz(float(toplam_tahsilat)))
        self.lblKalanBakiye.setText(para_yaz(float(kalan)))
        if hasattr(self, "lblDashboardBorc"):
            self.lblDashboardBorc.setText(para_yaz(float(kalan)))
        if hasattr(self, "lblDashboardTahsilat"):
            self.lblDashboardTahsilat.setText(para_yaz(float(toplam_tahsilat)))
        if hasattr(self, "dashboard_kpi_font_ayarla"):
            self.dashboard_kpi_font_ayarla()
        self.son_hareketleri_yukle()
        self.durum_cubugu_guncelle()

    def son_hareketleri_yukle(self):
        if not hasattr(self, "tblSonHareketler"):
            return

        conn = baglan()
        cur = conn.cursor()
        cur.execute("""
            SELECT h.tarih, c.ad, h.tip, h.tutar
            FROM hareketler h
            LEFT JOIN cariler c ON c.id = h.cari_id
            ORDER BY h.id DESC
            LIMIT 5
        """)
        hareketler = cur.fetchall()
        conn.close()

        self.tblSonHareketler.setRowCount(len(hareketler))
        self.tblSonHareketler.setVisible(bool(hareketler))
        if hasattr(self, "bosHareketPanel"):
            self.bosHareketPanel.setVisible(not bool(hareketler))

        for satir, (tarih, cari_ad, tip, tutar) in enumerate(hareketler):
            veriler = [str(tarih or ""), str(cari_ad or "-"), str(tip or ""), para_yaz(float(tutar or 0))]
            for sutun, alan in enumerate(veriler):
                item = QTableWidgetItem(str(alan))
                if tip == "BORÇ":
                    item.setData(Qt.UserRole, "BORÇ")
                elif tip == "TAHSİLAT":
                    item.setData(Qt.UserRole, "TAHSİLAT")
                self.tblSonHareketler.setItem(satir, sutun, item)

        if hasattr(self, "dashboard_ek_bilgileri_yukle"):
            self.dashboard_ek_bilgileri_yukle()
            # V149: Grafik çizimini UI döngüsünden sonra çalıştır; sayfa geçişi takılmasın.
            try:
                QTimer.singleShot(80, self.satis_grafik_guncelle)
            except Exception:
                pass

    def tarih_aralik_raporu(self):
        pencere = QDialog(self)
        pencere.setWindowTitle("Tarih Aralıklı Rapor")
        pencere.resize(950, 650)

        layout = QVBoxLayout()

        baslik = QLabel("TARİH ARALIKLI RAPOR")
        baslik.setStyleSheet("font-size:24px;font-weight:bold;padding:10px;color:#0D47A1;")
        layout.addWidget(baslik)

        filtre_layout = QHBoxLayout()

        filtre_layout.addWidget(QLabel("Başlangıç Tarihi"))
        dtBaslangic = QDateEdit()
        dtBaslangic.setCalendarPopup(True)
        bugun = QDate.currentDate()
        dtBaslangic.setDate(QDate(bugun.year(), bugun.month(), 1))
        filtre_layout.addWidget(dtBaslangic)

        filtre_layout.addWidget(QLabel("Bitiş Tarihi"))
        dtBitis = QDateEdit()
        dtBitis.setCalendarPopup(True)
        dtBitis.setDate(bugun)
        filtre_layout.addWidget(dtBitis)

        btnGetir = QPushButton("Raporu Getir")
        filtre_layout.addWidget(btnGetir)

        layout.addLayout(filtre_layout)

        ozet = QLabel("Rapor için tarih seçip 'Raporu Getir' butonuna basın.")
        ozet.setStyleSheet("font-size:16px;font-weight:bold;padding:10px;background:white;border-radius:8px;")
        layout.addWidget(ozet)

        tablo = QTableWidget()
        tablo.setColumnCount(6)
        tablo.setHorizontalHeaderLabels(["Cari", "Tarih", "Tip", "Tutar", "Açıklama", "Bakiye Etkisi"])
        tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tablo.setSelectionBehavior(QTableWidget.SelectRows)
        tablo.setEditTriggers(QTableWidget.NoEditTriggers)
        tablo.verticalHeader().setVisible(False)
        layout.addWidget(tablo)

        def raporu_getir():
            baslangic = dtBaslangic.date().toPython()
            bitis = dtBitis.date().toPython()

            if baslangic > bitis:
                QMessageBox.warning(pencere, "Hata", "Başlangıç tarihi bitiş tarihinden büyük olamaz.")
                return

            conn = baglan()
            cur = conn.cursor()
            cur.execute("""
                SELECT c.ad, h.tarih, h.tip, h.tutar, h.aciklama
                FROM hareketler h
                LEFT JOIN cariler c ON c.id = h.cari_id
                ORDER BY h.id
            """)
            hareketler = cur.fetchall()
            conn.close()

            filtreli = []
            toplam_borc = 0.0
            toplam_tahsilat = 0.0

            for cari_ad, tarih_text, tip, tutar, aciklama in hareketler:
                try:
                    islem_tarihi = datetime.strptime(str(tarih_text), "%d.%m.%Y %H:%M").date()
                except Exception:
                    continue

                if baslangic <= islem_tarihi <= bitis:
                    tutar = float(tutar or 0)

                    if tip == "BORÇ":
                        toplam_borc += tutar
                        etki = tutar
                    else:
                        toplam_tahsilat += tutar
                        etki = -tutar

                    filtreli.append([
                        cari_ad or "",
                        tarih_text,
                        tip,
                        para_yaz(tutar),
                        aciklama or "",
                        para_yaz(etki)
                    ])

            tablo.setRowCount(len(filtreli))

            for satir, veri in enumerate(filtreli):
                for sutun, alan in enumerate(veri):
                    tablo.setItem(satir, sutun, QTableWidgetItem(str(alan)))

            kalan = toplam_borc - toplam_tahsilat
            ozet.setText(
                f"Tarih Aralığı: {baslangic.strftime('%d.%m.%Y')} - {bitis.strftime('%d.%m.%Y')}\n"
                f"Toplam Borç: {para_yaz(toplam_borc)}    |    "
                f"Toplam Tahsilat: {para_yaz(toplam_tahsilat)}    |    "
                f"Kalan Etki: {para_yaz(kalan)}    |    "
                f"İşlem Sayısı: {len(filtreli)}"
            )

        btnGetir.clicked.connect(raporu_getir)

        btnKapat = QPushButton("Kapat")
        btnKapat.clicked.connect(pencere.close)
        layout.addWidget(btnKapat)

        pencere.setLayout(layout)
        pencere.exec()

    def raporlar_penceresi(self):
        pencere = QDialog(self)
        pencere.setWindowTitle("Raporlar")
        pencere.resize(900, 600)

        layout = QVBoxLayout()

        baslik = QLabel("RAPORLAR")
        baslik.setStyleSheet("font-size:24px;font-weight:bold;padding:10px;")
        layout.addWidget(baslik)

        conn = baglan()
        cur = conn.cursor()

        cur.execute("SELECT COUNT(*) FROM cariler WHERE COALESCE(aktif,1)=1")
        toplam_cari = cur.fetchone()[0] or 0

        cur.execute("SELECT SUM(tutar) FROM hareketler WHERE tip='BORÇ'")
        toplam_borc = cur.fetchone()[0] or 0

        cur.execute("SELECT SUM(tutar) FROM hareketler WHERE tip='TAHSİLAT'")
        toplam_tahsilat = cur.fetchone()[0] or 0

        kalan = float(toplam_borc) - float(toplam_tahsilat)

        ozet = QLabel(
            f"Toplam Cari: {toplam_cari}\n"
            f"Toplam Borç: {para_yaz(float(toplam_borc))}\n"
            f"Toplam Tahsilat: {para_yaz(float(toplam_tahsilat))}\n"
            f"Kalan Bakiye: {para_yaz(kalan)}"
        )
        ozet.setStyleSheet("font-size:18px;font-weight:bold;padding:12px;background:white;border-radius:8px;")
        layout.addWidget(ozet)

        tablo = QTableWidget()
        tablo.setColumnCount(5)
        tablo.setHorizontalHeaderLabels(["Cari", "Telefon", "Toplam Borç", "Toplam Tahsilat", "Bakiye"])
        tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tablo.setSelectionBehavior(QTableWidget.SelectRows)
        tablo.setEditTriggers(QTableWidget.NoEditTriggers)
        tablo.verticalHeader().setVisible(False)

        cur.execute("""
            SELECT
                c.ad,
                c.telefon,
                COALESCE(SUM(CASE WHEN h.tip='BORÇ' THEN h.tutar ELSE 0 END), 0) AS toplam_borc,
                COALESCE(SUM(CASE WHEN h.tip='TAHSİLAT' THEN h.tutar ELSE 0 END), 0) AS toplam_tahsilat
            FROM cariler c
                            LEFT JOIN hareketler h ON h.cari_id = c.id
            GROUP BY c.id, c.ad, c.telefon
            ORDER BY c.ad
        """)

        satirlar = cur.fetchall()
        conn.close()

        tablo.setRowCount(len(satirlar))

        for satir, veri in enumerate(satirlar):
            ad, telefon, borc, tahsilat = veri
            bakiye = float(borc or 0) - float(tahsilat or 0)

            bilgiler = [
                ad,
                telefon or "",
                para_yaz(float(borc or 0)),
                para_yaz(float(tahsilat or 0)),
                para_yaz(bakiye)
            ]

            for sutun, alan in enumerate(bilgiler):
                tablo.setItem(satir, sutun, QTableWidgetItem(str(alan)))

        layout.addWidget(tablo)

        btnKapat = QPushButton("Kapat")
        btnKapat.clicked.connect(pencere.close)
        layout.addWidget(btnKapat)

        pencere.setLayout(layout)
        pencere.exec()

    def excel_aktar(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            QMessageBox.warning(
                self,
                "Eksik Paket",
                "Excel aktarımı için openpyxl gerekli.\n\nTerminale şunu yaz:\npip install openpyxl"
            )
            return

        dosya_yolu, _ = QFileDialog.getSaveFileName(
            self,
            "Excel Dosyası Kaydet",
            "CariTakip_Yedek.xlsx",
            "Excel Dosyası (*.xlsx)"
        )

        if not dosya_yolu:
            return

        if not dosya_yolu.lower().endswith(".xlsx"):
            dosya_yolu += ".xlsx"

        conn = baglan()
        cur = conn.cursor()

        cur.execute("SELECT id, ad, telefon, adres, vergi_dairesi, vergi_no FROM cariler WHERE COALESCE(aktif,1)=1 ORDER BY ad")
        cariler = cur.fetchall()

        cur.execute("""
            SELECT h.id, c.ad, h.tarih, h.tip, h.tutar, h.aciklama
            FROM hareketler h
            LEFT JOIN cariler c ON c.id = h.cari_id
            ORDER BY h.id
        """)
        hareketler = cur.fetchall()

        conn.close()

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Cariler"

        basliklar = ["ID", "Ad Soyad", "Telefon", "Adres", "Vergi Dairesi", "Vergi No"]
        ws.append(basliklar)

        for cari in cariler:
            ws.append(cari)

        ws2 = wb.create_sheet("Hareketler")
        ws2.append(["ID", "Cari", "Tarih", "Tip", "Tutar", "Açıklama"])

        for hareket in hareketler:
            ws2.append(hareket)

        ws3 = wb.create_sheet("Özet")

        toplam_borc = sum(float(h[4] or 0) for h in hareketler if h[3] == "BORÇ")
        toplam_tahsilat = sum(float(h[4] or 0) for h in hareketler if h[3] == "TAHSİLAT")
        kalan = toplam_borc - toplam_tahsilat

        ws3.append(["Bilgi", "Değer"])
        firma = firma_bilgisi_getir()
        ws3.append(["Firma Adı", firma["firma_adi"]])
        ws3.append(["Firma Telefon", firma["telefon"]])
        ws3.append(["Firma Adres", firma["adres"]])
        ws3.append(["Firma Vergi No", firma["vergi_no"]])
        ws3.append(["Firma Vergi Dairesi", firma["vergi_dairesi"]])
        ws3.append(["Firma E-Posta", firma["eposta"]])
        ws3.append([])
        ws3.append(["Toplam Cari", len(cariler)])
        ws3.append(["Toplam Borç", toplam_borc])
        ws3.append(["Toplam Tahsilat", toplam_tahsilat])
        ws3.append(["Kalan Bakiye", kalan])

        for sayfa in [ws, ws2, ws3]:
            for cell in sayfa[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1976D2")
                cell.alignment = Alignment(horizontal="center")

            for col in sayfa.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    value = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(value))
                sayfa.column_dimensions[col_letter].width = max_len + 3

        try:
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", f"Excel dosyası oluşturuldu:\n{dosya_yolu}")
        except PermissionError:
            QMessageBox.warning(self, "Hata", "Excel dosyası açık olabilir. Kapatıp tekrar deneyin.")

    def tahsilat_makbuzu_penceresi(self):
        urun_tablolari_olustur()
        pencere = QDialog(self)
        pencere.setWindowTitle("Tahsilat Makbuzu")
        pencere.resize(900, 620)
        pencere.setStyleSheet("""
            QDialog { background:#EEF3F8; }
            QFrame#Panel { background:white;border:1px solid #D8E0EA;border-radius:14px; }
            QPushButton { background:#0D47A1;color:white;border:none;border-radius:10px;padding:9px;font-weight:bold; }
            QPushButton:hover { background:#1565C0; }
            QPushButton#GreenButton { background:#15803D; }
            QLineEdit, QTextEdit { background:white;border:1px solid #CBD5E1;border-radius:8px;padding:8px; }
            QTableWidget { background:white;border:1px solid #D8E0EA;border-radius:8px; }
        """)

        ana = QVBoxLayout()
        ana.setContentsMargins(12, 12, 12, 12)
        baslik = QLabel("TAHSİLAT MAKBUZU")
        baslik.setStyleSheet("font-size:24px;font-weight:bold;color:#0D47A1;")
        ana.addWidget(baslik)

        govde = QHBoxLayout()
        sol_panel = QFrame()
        sol_panel.setObjectName("Panel")
        sol = QVBoxLayout()
        sol.setContentsMargins(10, 10, 10, 10)
        sol.addWidget(QLabel("Cari Ara"))
        txtAra = QLineEdit()
        txtAra.setPlaceholderText("Müşteri adı / telefon ara...")
        sol.addWidget(txtAra)
        tabloCari = QTableWidget()
        tabloCari.setColumnCount(4)
        tabloCari.setHorizontalHeaderLabels(["No", "Cari", "Telefon", "ID"])
        tabloCari.setColumnHidden(3, True)
        tabloCari.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tabloCari.setSelectionBehavior(QTableWidget.SelectRows)
        tabloCari.setEditTriggers(QTableWidget.NoEditTriggers)
        tabloCari.verticalHeader().setVisible(False)
        sol.addWidget(tabloCari, 1)
        sol_panel.setLayout(sol)
        govde.addWidget(sol_panel, 4)

        sag_panel = QFrame()
        sag_panel.setObjectName("Panel")
        sag = QVBoxLayout()
        sag.setContentsMargins(14, 14, 14, 14)
        lblCari = QLabel("Seçili Cari: Yok")
        lblCari.setStyleSheet("font-size:17px;font-weight:bold;color:#0D47A1;background:#F3F4F6;border:1px solid #E5E7EB;border-radius:8px;padding:10px;")
        lblCari.setWordWrap(True)
        sag.addWidget(lblCari)
        sag.addWidget(QLabel("Tahsilat Tutarı"))
        txtTutar = QLineEdit()
        txtTutar.setPlaceholderText("Örn: 1500 veya 1.500,50")
        sag.addWidget(txtTutar)
        sag.addWidget(QLabel("Açıklama"))
        txtAciklama = QTextEdit()
        txtAciklama.setPlaceholderText("Örn: Proforma tahsilatı / peşinat / cari tahsilat")
        txtAciklama.setMaximumHeight(120)
        sag.addWidget(txtAciklama)
        btnKaydet = QPushButton("TAHSİLAT MAKBUZU OLUŞTUR")
        btnKaydet.setObjectName("GreenButton")
        sag.addWidget(btnKaydet)
        btnKapat = QPushButton("Kapat")
        btnKapat.clicked.connect(pencere.close)
        sag.addWidget(btnKapat)
        sag.addStretch()
        sag_panel.setLayout(sag)
        govde.addWidget(sag_panel, 3)
        ana.addLayout(govde)

        durum = {"cari": None}

        def cari_listele():
            arama = txtAra.text().strip()
            conn = baglan()
            cur = conn.cursor()
            if arama:
                like = f"%{arama}%"
                cur.execute("SELECT id, ad, telefon FROM cariler WHERE COALESCE(aktif,1)=1 AND (ad LIKE ? OR telefon LIKE ?) ORDER BY ad", (like, like))
            else:
                cur.execute("SELECT id, ad, telefon FROM cariler WHERE COALESCE(aktif,1)=1 ORDER BY ad")
            rows = cur.fetchall()
            conn.close()
            tabloCari.setRowCount(len(rows))
            for r, (cid, ad, tel) in enumerate(rows):
                tabloCari.setItem(r, 0, QTableWidgetItem(str(r + 1)))
                tabloCari.setItem(r, 1, QTableWidgetItem(ad or ""))
                tabloCari.setItem(r, 2, QTableWidgetItem(tel or ""))
                tabloCari.setItem(r, 3, QTableWidgetItem(str(cid)))

        def cari_sec():
            row = tabloCari.currentRow()
            if row < 0:
                return
            cid = int(tabloCari.item(row, 3).text())
            cari = self.cari_bilgisi_getir_id(cid)
            durum["cari"] = cari
            lblCari.setText(f"Seçili Cari: {cari['ad']}\nTel: {cari.get('telefon','')}")

        def makbuz_html(makbuz_id, cari, tutar, aciklama, tarih):
            firma = firma_bilgisi_getir()
            logo_yolu = os.path.join(BASE_DIR, "logo.png")
            logo_html = ""
            if os.path.exists(logo_yolu):
                logo_html = f'<img src="file:///{logo_yolu.replace(os.sep, "/")}" width="95" height="45">'

            firma_adi = firma.get('firma_adi', 'DAL ELEKTRONİK VE TEDARİK') or 'DAL ELEKTRONİK VE TEDARİK'
            telefon = firma.get('telefon', '') or ''
            vergi_no = firma.get('vergi_no', '') or ''
            adres = firma.get('adres', '') or ''
            eposta = firma.get('eposta', '') or ''

            return f"""
            <html>
            <head>
                <style>
                    body {{ font-family: Arial; font-size: 8.5pt; color:#111827; margin:0; }}
                    .firma {{ font-size: 8pt; line-height: 1.25; }}
                    .baslik {{ font-size: 17pt; font-weight: bold; color:#0D47A1; text-align:center; margin:7px 0 5px 0; }}
                    .kutu {{ border:1px solid #9CA3AF; padding:5px; margin-top:5px; }}
                    .bolum {{ font-weight:bold; color:#0D47A1; margin-bottom:3px; }}
                    table {{ border-collapse: collapse; }}
                    th {{ background:#0D47A1; color:white; font-weight:bold; padding:5px; }}
                    td {{ padding:5px; }}
                    .imza td {{ height:45px; vertical-align:top; }}
                </style>
            </head>
            <body>
                <table width='100%'>
                    <tr>
                        <td width='27%' valign='top'>{logo_html}</td>
                        <td width='73%' align='right' valign='top' class='firma'>
                            <b style='font-size:12pt;color:#0D47A1;'>{firma_adi}</b><br>
                            <b>Tel:</b> {telefon} &nbsp
                            | &nbsp
                            <b>Vergi No:</b> {vergi_no}<br>
                            <b>Adres:</b> {adres}<br>
                            <b>E-Posta:</b> {eposta}
                        </td>
                    </tr>
                </table>
                <hr>
                <div class='baslik'>TAHSİLAT MAKBUZU</div>
                <table width='100%'>
                    <tr>
                        <td><b>Makbuz No:</b> MK-{int(makbuz_id):05d}</td>
                        <td align='right'><b>Tarih:</b> {tarih}</td>
                    </tr>
                </table>

                <div class='kutu'>
                    <div class='bolum'>Müşteri Bilgileri</div>
                    <table width='100%'>
                        <tr><td width='22%'><b>Cari:</b></td><td>{cari.get('ad','')}</td></tr>
                        <tr><td><b>Telefon:</b></td><td>{cari.get('telefon','')}</td></tr>
                        <tr><td><b>Adres:</b></td><td>{cari.get('adres','')}</td></tr>
                        <tr><td><b>Vergi Dairesi:</b></td><td>{cari.get('vergi_dairesi','')}</td></tr>
                        <tr><td><b>Vergi No / T.C. No:</b></td><td>{cari.get('vergi_no','')}</td></tr>
                    </table>
                </div>

                <table width='100%' border='1' cellspacing='0' cellpadding='0' style='margin-top:7px;'>
                    <tr><th width='70%'>Açıklama</th><th width='30%'>Tutar</th></tr>
                    <tr><td>{aciklama or 'Cari tahsilat'}</td><td align='right'><b>{para_yaz(tutar)}</b></td></tr>
                </table>

                <div class='kutu' style='font-size:10pt; text-align:right;'>
                    <b>Tahsil Edilen Tutar: {para_yaz(tutar)}</b>
                </div>

                <table width='100%' class='imza' style='margin-top:9px;'>
                    <tr>
                        <td width='50%'><b>Tahsil Eden</b><br><br>İsim / Kaşe<br>İmza</td>
                        <td width='50%' align='right'><b>Ödeyen</b><br><br>Ad Soyad<br>İmza</td>
                    </tr>
                </table>
            </body>
            </html>
            """

        def kaydet():
            cari = durum.get("cari")
            if not cari:
                QMessageBox.warning(pencere, "Uyarı", "Önce cari seçin.")
                return
            tutar = self._sayi_oku(txtTutar.text())
            if tutar <= 0:
                QMessageBox.warning(pencere, "Uyarı", "Tahsilat tutarı 0'dan büyük olmalı.")
                return
            tarih = datetime.now().strftime("%d.%m.%Y %H:%M")
            aciklama = txtAciklama.toPlainText().strip() or "Tahsilat makbuzu"
            with db_baglan() as conn:
                cur = conn.cursor()
                cur.execute("INSERT INTO hareketler(cari_id, tip, tutar, aciklama, tarih) VALUES (?, 'TAHSİLAT', ?, ?, ?)", (cari["id"], tutar, aciklama, tarih))
                hareket_id = cur.lastrowid
                cur.execute("INSERT INTO tahsilat_makbuzlari(cari_id, tarih, tutar, aciklama, hareket_id) VALUES (?, ?, ?, ?, ?)", (cari["id"], tarih, tutar, aciklama, hareket_id))
                makbuz_id = cur.lastrowid
            klasor = os.path.join(BASE_DIR, "TahsilatMakbuzlari")
            os.makedirs(klasor, exist_ok=True)
            pdf_yolu = os.path.join(klasor, f"MK-{int(makbuz_id):05d}.pdf")
            try:
                printer = QPrinter(QPrinter.HighResolution)
                printer.setOutputFormat(QPrinter.PdfFormat)
                try:
                    printer.setPageSize(QPageSize(QPageSize.A5))
                except Exception:
                    pass
                printer.setOutputFileName(pdf_yolu)
                doc = QTextDocument()
                doc.setHtml(makbuz_html(makbuz_id, cari, tutar, aciklama, tarih))
                try:
                    doc.setPageSize(printer.pageRect(QPrinter.Point).size())
                except Exception:
                    pass
                doc.print_(printer)
                with db_baglan() as conn:
                    conn.cursor().execute("UPDATE tahsilat_makbuzlari SET pdf_yolu=? WHERE id=?", (pdf_yolu, makbuz_id))
            except Exception as hata:
                QMessageBox.warning(pencere, "PDF Hatası", f"Makbuz kaydedildi ama PDF oluşturulamadı:\n{hata}")
            self.ozet_yukle()
            QMessageBox.information(pencere, "Tahsilat Kaydedildi", f"Tahsilat makbuzu oluşturuldu.\nTutar: {para_yaz(tutar)}\n\nPDF:\n{pdf_yolu}")
            try:
                os.startfile(klasor)
            except Exception:
                pass
            if QMessageBox.question(pencere, "WhatsApp", "Makbuzu WhatsApp Web ile göndermek ister misiniz?") == QMessageBox.Yes:
                mesaj = f"Merhaba, {para_yaz(tutar)} tutarındaki tahsilat makbuzunuz hazırlanmıştır.\nDAL ELEKTRONİK VE TEDARİK"
                self._whatsapp_ac(cari.get("telefon", ""), mesaj)

        txtAra.textChanged.connect(cari_listele)
        tabloCari.cellClicked.connect(lambda r, c: cari_sec())
        tabloCari.cellDoubleClicked.connect(lambda r, c: cari_sec())
        btnKaydet.clicked.connect(kaydet)
        cari_listele()
        pencere.setLayout(ana)
        pencere.exec()


